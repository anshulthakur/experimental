from openpyxl import load_workbook

wb = load_workbook('Expenditure.xlsx')

#print wb.get_sheet_names()
ws = wb.active

#ws = wb._active_sheet_index
#ws  = create_sheet('Name', index_of_sheet)
#ws.title
#ws.sheet_properties.tabColor = "1072BA"

#Access worksheet by index: ws = wb["Name"]

#All sheet names: wb.sheetnames : it is a list
#`for sheet in wb` also works

#wb.copy_worksheet(wb.active)

#Access contents:
#ws['A4'] = 4
#     or
#ws.cell(row=4, column=2, value=4)

#range: ws['A1':'C2']
# ws['C']
# ws['C:D']
#row10 = ws[10]
#ws[5:10]

#   or

# for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
#    for cell in row:

#  or

# for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
#    for cell in col:

# ws.rows or ws.columns

# cell.value

